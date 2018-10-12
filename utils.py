# -*- coding: utf-8 -*-
"""
Created on Thu Oct 11 18:38:42 2018

@author: Reshma
"""

# -*- coding: utf-8 -*-
"""
Created on Thu Oct 11 18:58:01 2018

@author: Reshma
"""

import sqlite3
from sqlite3 import Error
import random 
from ecdsa import SigningKey, VerifyingKey, NIST384p

from openpyxl import Workbook
wb = Workbook()

def create_db(db_file):
    """ create a database connection to a SQLite database """
    try:
        conn = sqlite3.connect(db_file)
    except Error as e:
        print(e)
        
    
    conn.execute("DROP TABLE TRXNS")
    conn.execute('''CREATE TABLE TRXNS
             (ID INT PRIMARY KEY     NOT NULL,
             TO_ACC           INT    NOT NULL,
             FROM_ACC            INT     NOT NULL,
             STAMP       INT NOT NULL,
             AMT         REAL);''')
    
    #print("Table created successfully")
    conn.commit()
    conn.close()

def ins_db(db_file):    
    
    conn = sqlite3.connect(db_file)
    #print("Opened database successfully")
    
    for i in range(0,10):
        ID=i
        TO_ACC = 100000000 + random.randint(0,100000000)
        FROM_ACC = 100000000 + random.randint(0,100000000)
        STAMP = i+1028
        AMT = random.randint(0,100000)
        conn.execute("""Insert into TRXNS values(?,?,?,?,?) """,(ID,TO_ACC,FROM_ACC,STAMP,AMT));
        
    cur = conn.execute("SELECT * from TRXNS")
    r = cur.fetchall()
    conn.commit()
    conn.close()
    return r


def disp_db(db_file):
    
    conn = sqlite3.connect(db_file)
    cur = conn.cursor()
    cur.execute("SELECT * from TRXNS")
    rows = cur.fetchall()
    for row in rows:
       print(row)
      
    #print("Operation done successfully")
    conn.close()

import xlwt
from xlwt import Workbook   

def read_wb(name):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.create_sheet(name)
    
def write_wb(name , data):
    ws = wb.create_sheet(name)
    for i in data :
        for j in i:
            ws.write(j)
            ws.append((i))#row=i, column=j) = 1#data[j][j]
    #ws.append(data)
    wb.save(name)
        
def disp_ws(ws):
    for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
        for cell in row:
            print(cell) 

def sql_source(fname):
    create_db(fname)
    conn = sqlite3.connect(fname)
    cur = conn.cursor()
    cur.execute("SELECT * from TRXNS")
    r = cur.fetchall()
    conn.close()
    return r

def rnd_source(fname):
    #create_db(fname)
    return ins_db(fname)

def excel_source(fname):
    read_wb(fname)
    

    